"""
Generadores de PDF y Excel para Proyecto Córdoba.
PDF: xhtml2pdf (puro Python, sin libpango/cairo).
Excel: openpyxl.
Principio de privacidad: solo código de paciente, nunca nombre real.
"""
import base64
import logging
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from collections import defaultdict

from django.template.loader import render_to_string
from django.utils import timezone

logger = logging.getLogger(__name__)

REPORT_STATUSES = ('approved', 'settled', 'exported')
CATEGORY_LABELS = {
    'transport': 'Transporte',
    'meals': 'Comidas',
    'accommodation': 'Alojamiento',
    'other': 'Otro',
}


def _ticket_to_base64(ticket) -> str | None:
    """Convierte el archivo del ticket a data URI base64 para incrustar en el PDF."""
    if not ticket or not ticket.file:
        return None
    try:
        mime = ticket.mime_type or 'image/jpeg'
        if 'pdf' in mime:
            return None  # No se incrusta PDF en el PDF
        with ticket.file.open('rb') as f:
            data = f.read()
        return f"data:{mime};base64,{base64.b64encode(data).decode()}"
    except Exception as e:
        logger.warning("No se pudo convertir ticket a base64: %s", e)
        return None


def _calc_totals(expenses) -> dict:
    """
    Calcula totales por categoría y total general, SIEMPRE separados por moneda.
    Nunca se suman montos de monedas distintas en un mismo total.

    Estructura:
    - by_currency: lista de bloques {currency, by_category, total} ordenada
      por relevancia (la moneda con más gastos primero).
    - by_category / total / currency: alias del bloque dominante, para
      compatibilidad con templates que asumen una única moneda.
    - mixed: True si hay más de una moneda en el conjunto.
    """
    per_currency: dict = {}
    for exp in expenses:
        cur = exp.currency or 'ARS'
        block = per_currency.setdefault(cur, {'by_category': defaultdict(Decimal), 'count': 0})
        block['by_category'][exp.category] += exp.amount
        block['count'] += 1

    by_currency = []
    for cur, block in sorted(per_currency.items(), key=lambda kv: -kv[1]['count']):
        cats = {
            cat: {'label': CATEGORY_LABELS.get(cat, cat), 'amount': amt}
            for cat, amt in sorted(block['by_category'].items())
        }
        by_currency.append({
            'currency': cur,
            'by_category': cats,
            'total': sum(block['by_category'].values()),
        })

    dominant = by_currency[0] if by_currency else {
        'currency': 'ARS', 'by_category': {}, 'total': Decimal('0'),
    }
    return {
        'by_currency': by_currency,
        'by_category': dominant['by_category'],
        'total': dominant['total'],
        'currency': dominant['currency'],
        'mixed': len(by_currency) > 1,
    }


def _render_pdf(template_name: str, context: dict) -> bytes:
    """Renderiza un template HTML a PDF usando xhtml2pdf."""
    from xhtml2pdf import pisa

    html = render_to_string(template_name, context)
    buffer = BytesIO()
    status = pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    if status.err:
        raise RuntimeError(f"Error xhtml2pdf ({status.err}): verifica el template {template_name}")
    return buffer.getvalue()


def _mark_expenses_exported(expenses, user, report_type: str, report_label: str):
    """
    Marca los gastos como 'exported' y crea un AuditLog por cada uno.
    Idempotente: no crea duplicados si el gasto ya estaba exported.
    """
    from apps.expenses.models import AuditLog

    expense_ids = [e.pk for e in expenses]
    AuditLog.objects.bulk_create([
        AuditLog(
            user=user,
            action='exported',
            content_type='Expense',
            object_id=exp.pk,
            object_repr=str(exp),
            details={
                'report_type': report_type,
                'report_label': report_label,
                'prev_status': exp.status,
            },
        )
        for exp in expenses
    ])
    # Actualizar status en bulk — incluye approved y settled, no toca los ya exportados
    from apps.expenses.models import Expense
    Expense.objects.filter(
        pk__in=expense_ids, status__in=('approved', 'settled')
    ).update(status='exported')


# ─── PDF por paciente ──────────────────────────────────────────────────────────

def generate_patient_pdf(patient, period, requested_by) -> bytes:
    """
    Genera el PDF de rendición de un paciente en un período.
    Solo incluye gastos con status en REPORT_STATUSES.
    Incrusta miniaturas de tickets como base64.
    Marca los gastos como 'exported' y crea AuditLog.
    """
    from apps.expenses.models import Expense

    expenses = list(
        Expense.objects.filter(
            visit__patient=patient,
            expense_date__gte=period.date_from,
            expense_date__lte=period.date_to,
            status__in=REPORT_STATUSES,
        ).select_related(
            'visit__visit_type',
        ).prefetch_related('ticket_files')
        .order_by('expense_date', 'visit__visit_type__order')
    )

    if not expenses:
        raise ValueError(
            f"No hay gastos aprobados para {patient.patient_code} en el período {period.name}."
        )

    # Preparar datos de gastos con imagen base64
    expense_rows = []
    tickets_with_images = []
    total_viaticos_usd = Decimal('0')

    for exp in expenses:
        ticket = exp.ticket_files.first()
        ticket_b64 = _ticket_to_base64(ticket)
        expense_rows.append({
            'expense': exp,
            'ticket_b64': ticket_b64,
            'ticket_filename': ticket.original_filename if ticket else '',
        })

        # Acumular para viáticos y galería
        total_viaticos_usd += exp.amount_usd or Decimal('0')
        if ticket_b64:
            tickets_with_images.append({
                'ticket_b64': ticket_b64,
                'date': exp.expense_date,
                'vendor': exp.vendor,
                'category': exp.get_category_display(),
                'amount': exp.amount,
                'currency': exp.currency,
            })

    totals = _calc_totals(expenses)

    # Calcular viáticos
    remaining_viaticos = Decimal('0')
    percentage_used = 0
    if patient.viatic_cap:
        remaining_viaticos = patient.viatic_cap - total_viaticos_usd
        percentage_used = int((total_viaticos_usd / patient.viatic_cap) * 100)

    context = {
        'patient': patient,
        'period': period,
        'protocol': patient.protocol,
        'expenses': expense_rows,
        'totals': totals,
        'total_viaticos_usd': total_viaticos_usd,
        'remaining_viaticos': remaining_viaticos,
        'percentage_used': percentage_used,
        'tickets_with_images': tickets_with_images,
        'generated_by': requested_by,
        'generated_at': timezone.now(),
        'report_type': 'patient',
    }

    pdf_bytes = _render_pdf('reports/pdf_patient.html', context)
    report_label = f"PDF paciente {patient.patient_code} / {patient.protocol.code} / {period.name}"
    _mark_expenses_exported(expenses, requested_by, 'patient_pdf', report_label)

    return pdf_bytes


# ─── PDF consolidado del site ─────────────────────────────────────────────────

def generate_site_pdf(protocol, period, requested_by) -> bytes:
    """
    Genera el PDF consolidado de todos los pacientes del protocolo en el período.
    Primera página: resumen ejecutivo con totales.
    Sección por paciente: detalle de gastos.
    """
    from apps.patients.models import Patient
    from apps.expenses.models import Expense

    patients = Patient.objects.filter(
        protocol=protocol,
        is_active=True,
    ).order_by('patient_code')

    all_expenses = []
    patient_sections = []

    for patient in patients:
        expenses = list(
            Expense.objects.filter(
                visit__patient=patient,
                expense_date__gte=period.date_from,
                expense_date__lte=period.date_to,
                status__in=REPORT_STATUSES,
            ).select_related('visit__visit_type')
            .prefetch_related('ticket_files')
            .order_by('expense_date')
        )
        if not expenses:
            continue

        expense_rows = []
        for exp in expenses:
            ticket = exp.ticket_files.first()
            expense_rows.append({
                'expense': exp,
                'ticket_b64': _ticket_to_base64(ticket),
                'ticket_filename': ticket.original_filename if ticket else '',
            })

        totals = _calc_totals(expenses)
        patient_sections.append({
            'patient': patient,
            'expenses': expense_rows,
            'totals': totals,
        })
        all_expenses.extend(expenses)

    if not all_expenses:
        raise ValueError(
            f"No hay gastos aprobados para ningún paciente de {protocol.code} en {period.name}."
        )

    # Resumen ejecutivo: totales por paciente y moneda (una fila por combinación)
    executive_summary = []
    grand_totals_by_currency: dict = {}
    for section in patient_sections:
        for block in section['totals']['by_currency']:
            executive_summary.append({
                'patient_code': section['patient'].patient_code,
                'currency': block['currency'],
                'totals_by_cat': block['by_category'],
                'total': block['total'],
            })
            grand_totals_by_currency[block['currency']] = (
                grand_totals_by_currency.get(block['currency'], Decimal('0')) + block['total']
            )

    grand_totals = [
        {'currency': cur, 'total': total}
        for cur, total in sorted(grand_totals_by_currency.items())
    ]

    context = {
        'protocol': protocol,
        'period': period,
        'patient_sections': patient_sections,
        'executive_summary': executive_summary,
        'grand_totals': grand_totals,
        'generated_by': requested_by,
        'generated_at': timezone.now(),
        'report_type': 'site',
    }

    pdf_bytes = _render_pdf('reports/pdf_site.html', context)
    report_label = f"PDF consolidado {protocol.code} / {period.name}"
    _mark_expenses_exported(all_expenses, requested_by, 'site_pdf', report_label)

    return pdf_bytes


# ─── Excel consolidado ────────────────────────────────────────────────────────

def generate_site_excel(protocol, period, requested_by) -> bytes:
    """
    Genera un archivo Excel con hoja de resumen y hoja por paciente.
    Usa openpyxl con estilos básicos.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from apps.patients.models import Patient
    from apps.expenses.models import Expense

    wb = openpyxl.Workbook()

    # Estilo de cabecera
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    HEADER_FILL = PatternFill('solid', fgColor='1E3A8A')  # primary-900
    CENTER = Alignment(horizontal='center', vertical='center')
    THIN = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(cell):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    def style_cell(cell):
        cell.border = BORDER
        cell.alignment = Alignment(vertical='center')

    # ── Hoja de resumen ejecutivo ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = 'Resumen'
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 10
    ws_summary.column_dimensions['C'].width = 16
    ws_summary.column_dimensions['D'].width = 16
    ws_summary.column_dimensions['E'].width = 16
    ws_summary.column_dimensions['F'].width = 16
    ws_summary.column_dimensions['G'].width = 16

    # Título
    ws_summary['A1'] = f'Proyecto Córdoba — Reporte Consolidado'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = f'Protocolo: {protocol.code} — {protocol.name}'
    ws_summary['A3'] = f'Período: {period.name} ({period.date_from} — {period.date_to})'
    ws_summary['A4'] = f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")} UTC'
    ws_summary['A4'].font = Font(italic=True, color='888888')

    ws_summary.row_dimensions[6].height = 18
    headers = ['Paciente', 'Moneda', 'Transporte', 'Comidas', 'Alojamiento', 'Otro', 'TOTAL']
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=6, column=col, value=h)
        style_header(cell)

    patients = Patient.objects.filter(protocol=protocol, is_active=True).order_by('patient_code')
    all_expenses = []
    row_num = 7
    # Totales generales por (moneda, categoría) — nunca se mezclan monedas.
    grand_totals: dict = defaultdict(lambda: defaultdict(Decimal))

    for patient in patients:
        expenses = list(
            Expense.objects.filter(
                visit__patient=patient,
                expense_date__gte=period.date_from,
                expense_date__lte=period.date_to,
                status__in=REPORT_STATUSES,
            ).select_related('visit__visit_type')
            .prefetch_related('ticket_files')
            .order_by('expense_date')
        )
        if not expenses:
            continue

        all_expenses.extend(expenses)
        totals = _calc_totals(expenses)
        for block in totals['by_currency']:
            cur = block['currency']
            row = [patient.patient_code, cur]
            for cat in ('transport', 'meals', 'accommodation', 'other'):
                amt = block['by_category'].get(cat, {}).get('amount', Decimal('0'))
                grand_totals[cur][cat] += amt
                row.append(float(amt))
            row.append(float(block['total']))

            for col, val in enumerate(row, 1):
                cell = ws_summary.cell(row=row_num, column=col, value=val)
                style_cell(cell)
                if col > 2:
                    cell.number_format = '#,##0.00'

            row_num += 1

    # Filas de totales generales, una por moneda
    for cur in sorted(grand_totals):
        grand_row = ['TOTAL GENERAL', cur]
        for cat in ('transport', 'meals', 'accommodation', 'other'):
            grand_row.append(float(grand_totals[cur][cat]))
        grand_row.append(float(sum(grand_totals[cur].values())))
        for col, val in enumerate(grand_row, 1):
            cell = ws_summary.cell(row=row_num, column=col, value=val)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='DBEAFE')
            cell.border = BORDER
            if col > 2:
                cell.number_format = '#,##0.00'
        row_num += 1

    if not all_expenses:
        raise ValueError(
            f"No hay gastos aprobados para {protocol.code} en {period.name}."
        )

    # ── Hojas por paciente ─────────────────────────────────────────────────────
    patients_with_expenses = Patient.objects.filter(
        protocol=protocol,
        is_active=True,
        visits__expenses__expense_date__gte=period.date_from,
        visits__expenses__expense_date__lte=period.date_to,
        visits__expenses__status__in=REPORT_STATUSES,
    ).distinct().order_by('patient_code')

    expense_detail_headers = [
        'Fecha', 'Visita', 'Categoría', 'Proveedor', 'CUIT',
        'N° Comprobante', 'Monto', 'Moneda', 'Estado',
    ]

    for patient in patients_with_expenses:
        expenses = list(
            Expense.objects.filter(
                visit__patient=patient,
                expense_date__gte=period.date_from,
                expense_date__lte=period.date_to,
                status__in=REPORT_STATUSES,
            ).select_related('visit__visit_type')
            .order_by('expense_date')
        )
        if not expenses:
            continue

        sheet_name = patient.patient_code[:31]  # Excel max 31 chars
        ws = wb.create_sheet(title=sheet_name)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 14

        ws['A1'] = f'Paciente: {patient.patient_code}'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f'Protocolo: {protocol.code} — Período: {period.name}'

        for col, h in enumerate(expense_detail_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            style_header(cell)

        for row_i, exp in enumerate(expenses, 5):
            ocr = exp.ocr_extracted if hasattr(exp, 'ocr_extracted') else {}
            row_data = [
                exp.expense_date.strftime('%d/%m/%Y'),
                exp.visit.visit_type.name,
                CATEGORY_LABELS.get(exp.category, exp.category),
                exp.vendor or '',
                ocr.get('cuit', '') if ocr else '',
                ocr.get('receipt_number', '') if ocr else '',
                float(exp.amount),
                exp.currency,
                exp.get_status_display(),
            ]
            for col_i, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                style_cell(cell)
                if col_i == 7:
                    cell.number_format = '#,##0.00'

        # Filas de total, una por moneda presente en la hoja
        totals = _calc_totals(expenses)
        total_row = row_i + 1 if expenses else 5
        for block in totals['by_currency']:
            ws.cell(row=total_row, column=6, value=f"TOTAL {block['currency']}").font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=7, value=float(block['total']))
            total_cell.font = Font(bold=True)
            total_cell.number_format = '#,##0.00'
            total_cell.fill = PatternFill('solid', fgColor='DBEAFE')
            ws.cell(row=total_row, column=8, value=block['currency']).font = Font(bold=True)
            total_row += 1

    # Guardar y retornar bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    report_label = f"Excel consolidado {protocol.code} / {period.name}"
    _mark_expenses_exported(all_expenses, requested_by, 'site_excel', report_label)

    return buffer.getvalue()
